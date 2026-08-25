import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest

from src.scoring.scoring import compute_engagement, compute_scores
from src.utils.config import load_settings


@pytest.fixture(scope="module")
def settings():
    return load_settings()


def _base_candidate(**overrides):
    cand = {
        "california_relevance": "HIGH",
        "trucking_relevance": "HIGH",
        "activity_status": "ACTIVE",
        "followers": 5000,
        "bio": "CA trucking creator",
        "_engagement_ratio": 0.10,
        "engagement_status": "KNOWN",
        "public_contact": "business email in bio",
        "other_social_links": ["https://instagram.com/example"],
        "trucking_topics": ["CDL", "otr", "diesel"],
        "_source_count": 2,
    }
    cand.update(overrides)
    return cand


class TestScoring:
    def test_max_score_100(self, settings):
        score, conf, breakdown = compute_scores(_base_candidate(), settings)
        assert score == 100.0
        assert breakdown["breakdown"]["california"]["points"] == 25
        assert breakdown["breakdown"]["trucking"]["points"] == 25
        assert breakdown["breakdown"]["activity"]["points"] == 20
        assert breakdown["breakdown"]["audience_fit"]["points"] == 10
        assert breakdown["breakdown"]["engagement"]["points"] == 10
        assert breakdown["breakdown"]["marketing_potential"]["points"] == 10

    def test_missing_data_never_inflates(self, settings):
        cand = _base_candidate(
            california_relevance="UNKNOWN",
            trucking_relevance="UNKNOWN",
            activity_status="UNKNOWN",
            followers=None,
            _engagement_ratio=None,
            engagement_status="UNKNOWN",
            public_contact=None,
            other_social_links=[],
            trucking_topics=[],
        )
        score, conf, _ = compute_scores(cand, settings)
        assert score < 20
        assert conf < 50

    def test_unknown_followers_gets_no_fit_points(self, settings):
        _, _, b = compute_scores(_base_candidate(followers=None), settings)
        assert b["breakdown"]["audience_fit"]["points"] == 0

    def test_out_of_band_followers_no_fit(self, settings):
        _, _, b = compute_scores(_base_candidate(followers=50000), settings)
        assert b["breakdown"]["audience_fit"]["points"] == 0

    def test_confidence_tracks_verification(self, settings):
        full_conf = compute_scores(_base_candidate(), settings)[1]
        weak_cand = _base_candidate(
            followers=None, bio="", _engagement_ratio=None, engagement_status="UNKNOWN", _source_count=0
        )
        weak_conf = compute_scores(weak_cand, settings)[1]
        assert full_conf > weak_conf
        assert full_conf <= 100 and weak_conf >= 0


class TestEngagement:
    def test_known_ratio(self):
        ratio, status, detail = compute_engagement(likes=800, comments=200, views=10000)
        assert status == "KNOWN"
        assert abs(ratio - 0.1) < 1e-6

    def test_insufficient_no_views(self):
        ratio, status, _ = compute_engagement(likes=10, comments=5, views=0)
        assert status == "UNKNOWN"
        assert ratio is None

    def test_insufficient_no_interactions(self):
        ratio, status, _ = compute_engagement(likes=0, comments=0, views=1000)
        assert status == "UNKNOWN"


class TestDatabaseAndResume:
    def _db(self, tmp_path):
        from src.database.db import Database
        return Database(str(tmp_path / "t.sqlite3"))

    def test_run_lifecycle_and_resume_query(self, tmp_path):
        db = self._db(tmp_path)
        db.create_run("run_1", limit_candidates=5)
        run = db.get_run("run_1")
        assert run["status"] == "running"
        db.upsert_candidate_discovery("a", "https://www.tiktok.com/@a", run_id="run_1")
        db.upsert_candidate_discovery("b", "https://www.tiktok.com/@b", run_id="run_1")
        pending = db.pending_candidates("run_1")
        assert len(pending) == 2
        db.update_candidate(pending[0]["id"], verification_status="done", follower_status="QUALIFIED")
        still_pending = db.pending_candidates("run_1")
        assert len(still_pending) == 1
        db.finish_run("run_1", "incomplete")
        latest = db.latest_running_run()
        assert latest is None or latest["run_id"] != "run_1"
        db.close()

    def test_processed_url_skip(self, tmp_path):
        db = self._db(tmp_path)
        url = "https://www.tiktok.com/@done_already"
        assert not db.was_processed_ok(url)
        db.record_processed(url, "profile", "ok", "run_1")
        assert db.was_processed_ok(url)
        db.record_processed(url, "profile", "blocked", "run_1")
        assert not db.was_processed_ok(url)
        db.close()

    def test_error_recording(self, tmp_path):
        db = self._db(tmp_path)
        db.record_error("run_1", "tiktok_web", "https://x", "blocked", "captcha wall")
        errs = db.query("SELECT * FROM errors")
        assert len(errs) == 1
        assert errs[0]["error_type"] == "blocked"
        db.close()

    def test_seeds_persisted_once(self, tmp_path):
        from src.database.db import Database
        db = self._db(tmp_path)
        seeds = [{"name": "Alex Nino", "aliases": ["alexnino"], "platforms": ["tiktok"]}]
        db.ensure_seeds(seeds)
        db.ensure_seeds(seeds)
        n = db.one("SELECT COUNT(*) n FROM seed_accounts")["n"]
        assert n == 1
        db.close()

    def test_activity_and_score_history(self, tmp_path):
        db = self._db(tmp_path)
        cid, _ = db.upsert_candidate_discovery("c", "https://www.tiktok.com/@c")
        db.save_activity(cid, "2026-08-24T00:00:00+00:00", None, 5, "UNKNOWN", "no_date_available")
        db.save_score(cid, 42.0, 66.6, {"any": 1})
        a = db.query("SELECT * FROM candidate_activity WHERE candidate_id=?", (cid,))
        s = db.query("SELECT * FROM candidate_scores WHERE candidate_id=?", (cid,))
        assert len(a) == 1 and len(s) == 1
        db.close()


class TestExportFilters:
    def test_california_only_filter(self, settings):
        from src.exporters.exporters import apply_export_filters
        rows = [
            {"rank": 1, "username": "@high", "california_relevance": "HIGH", "marketing_score": 90},
            {"rank": 2, "username": "@mid", "california_relevance": "MEDIUM", "marketing_score": 80},
            {"rank": 3, "username": "@low", "california_relevance": "LOW", "marketing_score": 70},
            {"rank": 4, "username": "@unk", "california_relevance": "UNKNOWN", "marketing_score": 60},
        ]
        kept = apply_export_filters([dict(r) for r in rows], settings)
        assert [r["username"] for r in kept] == ["@high", "@mid"]
        assert [r["rank"] for r in kept] == [1, 2]

    def test_filter_disabled(self):
        from src.exporters.exporters import apply_export_filters
        cfg = {"export": {"california_only": False}}
        rows = [{"rank": 1, "username": "@low", "california_relevance": "LOW"}]
        assert len(apply_export_filters(rows, cfg)) == 1


class TestExports:
    def _seeded_db(self, tmp_path):
        from src.database.db import Database
        db = Database(str(tmp_path / "exp.sqlite3"))
        db.create_run("run_e", limit_candidates=3)
        for uname, followers, ca in [("aaa", 3000, "HIGH"), ("bbb", 8000, "MEDIUM"), ("ccc", 1500, "LOW")]:
            cid, _ = db.upsert_candidate_discovery(uname, f"https://www.tiktok.com/@{uname}", run_id="run_e")
            db.add_source(cid, "Big Rig Videos" if uname == "aaa" else None, "search_engine")
            fields = {
                "display_name": uname.upper(),
                "followers": followers,
                "follower_status": ("QUALIFIED" if 2000 <= followers < 10000 else ("TOO_SMALL" if followers < 2000 else "TOO_LARGE")),
                "activity_status": "ACTIVE",
                "last_post_date": "2026-08-01T00:00:00+00:00",
                "recent_post_date": "2026-08-01T00:00:00+00:00",
                "recent_post_count": 4,
                "trucking_relevance": "HIGH",
                "trucking_topics": ["CDL"],
                "california_relevance": ca,
                "california_evidence": "Fresno" if ca == "HIGH" else "",
                "engagement_status": "KNOWN",
                "engagement_indicator": "test",
                "date_checked": "2026-08-24T00:00:00+00:00",
                "verification_status": "done",
            }
            db.update_candidate(cid, **fields)
            score = {"HIGH": 25, "MEDIUM": 15, "LOW": 7}[ca] + 25 + 20 + (10 if 2000 <= followers < 10000 else 0) + 10 + 5
            db.save_score(cid, float(score), 80.0, {})
        return db

    def test_csv_export_headers_and_sorting(self, tmp_path, settings):
        from src.exporters.exporters import build_rows, export_csv
        db = self._seeded_db(tmp_path)
        rows = build_rows(db)
        assert len(rows) == 2
        scores = [r["marketing_score"] for r in rows]
        assert scores == sorted(scores, reverse=True)
        top = rows[0]
        assert top["username"].startswith("@")
        assert top["all_seed_sources"] == "Big Rig Videos"
        path = export_csv(rows, {**settings, "export": {"csv_file": "test.csv"}, "app": {"output_dir": str(tmp_path / "out")}})
        import csv as csvmod
        with open(path, encoding="utf-8-sig") as f:
            reader = list(csvmod.DictReader(f))
        assert len(reader) == 2
        assert set(["rank", "username", "marketing_score"]).issubset(reader[0].keys())
        assert int(reader[0]["rank"]) == 1
        db.close()

    def test_xlsx_export_valid_file(self, tmp_path, settings):
        from openpyxl import load_workbook
        from src.exporters.exporters import build_rows, export_xlsx
        db = self._seeded_db(tmp_path)
        rows = build_rows(db)
        outdir = str(tmp_path / "out2")
        path = export_xlsx(rows, {"export": {"xlsx_file": "test.xlsx"}, "app": {"output_dir": outdir}})
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers[0] == "rank"
        assert ws.max_row == len(rows) + 1
        db.close()

    def test_json_export_structure(self, tmp_path, settings):
        import json
        from src.exporters.exporters import build_rows, export_json
        db = self._seeded_db(tmp_path)
        rows = build_rows(db)
        outdir = str(tmp_path / "out3")
        path = export_json(rows, {"export": {"json_file": "test.json"}, "app": {"output_dir": outdir}})
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        assert data["count"] == 2
        assert isinstance(data["creators"], list)
        assert data["creators"][0]["profile_url"].startswith("https://www.tiktok.com/@")
        db.close()

    def test_summary_report_content(self, tmp_path, settings):
        from src.exporters.exporters import build_summary
        db = self._seeded_db(tmp_path)
        text = build_summary(db, {**settings, "app": {"output_dir": str(tmp_path)}})
        assert "California Trucking TikTok Creator Research Summary" in text
        assert "Qualified (2k-10k followers): 2" in text
        assert "Top 25 Qualified Creators" in text
        assert "@aaa" in text
        db.close()
